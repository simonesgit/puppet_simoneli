import json
import requests
import openpyxl

# Jira project key and API URL
project_key = "YOUR_PROJECT_KEY"
api_url = "https://your-jira-instance.com/rest/api/2/issue"

# Define the path to the CA file
ca_file_path = "/path/to/your/ca_file.pem"

# Read Jira credentials from JSON file
with open('jira_credentials.json', 'r') as file:
    credentials = json.load(file)

username = credentials["username"]
password = credentials["password"]

def main():
    # Load data from Excel file
    workbook = openpyxl.load_workbook('input.xlsx')
    sheet = workbook.active

    # Get column names from the first row of the sheet
    header_row = sheet[1]
    column_names = [cell.value for cell in header_row]

    # Find the index of the issuekey column
    issuekey_index = column_names.index("issuekey")

    # Prepare Jira issue data for each row in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check if the issuekey is empty for the current row
        if not row[issuekey_index]:
            # Create a dictionary mapping column names to row values
            row_data = {column_names[i]: row[i] for i in range(len(column_names))}

            # Prepare Jira issue data
            issue_data = {
                "fields": {
                    "project": {"key": project_key},
                    "issuetype": {"name": row_data["issuetype"]},
                    "assignee": {"name": row_data["assignee"]},
                    # Add more fields as needed
                }
            }

            # Add additional fields from the Excel file to the issue_data
            for field_name, field_value in row_data.items():
                issue_data["fields"][field_name] = str(field_value)

            # Remove empty values from the issue data
            issue_data = {k: v for k, v in issue_data.items() if v}

            # Send POST request to create the issue
            response = requests.post(
                api_url,
                auth=(username, password),
                json=issue_data,
                verify=ca_file_path
            )

            # Check response status
            if response.status_code == 201:
                # Extract the issue key from the response
                new_issue_key = response.json()["key"]
                print("Issue created successfully:", new_issue_key)

                # Update the issuekey in the Excel file for the corresponding line
                sheet.cell(row=row[0].row, column=issuekey_index + 1, value=new_issue_key)
            else:
                print("Failed to create issue. Status code:", response.status_code)
                print("Error message:", response.text)

    # Save the updated Excel file
    workbook.save("input.xlsx")

if __name__ == "__main__":
    main()
