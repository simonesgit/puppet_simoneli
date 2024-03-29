import pandas as pd

# Step 1: Read contents from the three reports and store as dataframes
initiatives_df = pd.read_csv('export_all_initiatives.csv')
epics_df = pd.read_csv('export_all_epics.csv')
stories_df = pd.read_csv('export_all_stories.csv')

# Step 2: Rename column names with prefixes based on the source file
initiatives_df = initiatives_df.add_prefix('I_')
epics_df = epics_df.add_prefix('E_')
stories_df = stories_df.add_prefix('S_')

# Add the 'assignee' column information in each dataframe
# initiatives_df.rename(columns={'I_assignee': 'assignee'}, inplace=True)
# epics_df.rename(columns={'E_assignee': 'assignee'}, inplace=True)
# stories_df.rename(columns={'S_assignee': 'assignee'}, inplace=True)

# Step 3: Update target end if due date is later or there's no target date
initiatives_df['I_TargetEnd'] = initiatives_df.apply(lambda row: row['I_EndDate'] if pd.notnull(row['I_EndDate']) else row['I_TargetEnd'], axis=1)
epics_df['E_TargetEnd'] = epics_df.apply(lambda row: row['E_duedate'] if pd.notnull(row['E_duedate']) else row['E_TargetEnd'], axis=1)

# Step 4: Create 'resource' column for each dataframe
initiatives_df['resource'] = ''
epics_df['resource'] = ''
stories_df['resource'] = ''

# Step 5: Assign resource value of stories
stories_df['resource_list'] = stories_df.apply(lambda row: [row['S_assignee']] + str(row['S_AdditionalAssignee']).split(' | '), axis=1)
stories_df['resource'] = stories_df['resource_list'].apply(lambda x: ', '.join(list(set(filter(None, map(str, x))))))  # Convert values to strings, remove duplicates, and generate resource string

# Step 6: Assign resource value of epics
unique_epic_ids = epics_df['E_key'].unique()
for epic_id in unique_epic_ids:
    epic_row = epics_df[epics_df['E_key'] == epic_id].iloc[0]
    epic_stories = stories_df[stories_df['S_EpicLink'] == epic_id]
    epic_resource_list = [str(epic_row['E_assignee'])] + str(epic_row['E_AdditionalAssignee']).split(' | ') + list(epic_stories['S_assignee'].astype(str)) + list(epic_stories['S_AdditionalAssignee'].astype(str))
    epic_resource_list = [name for name in epic_resource_list if str(name) != 'nan']  # Remove 'nan' values
    epic_resource = ', '.join(list(set(filter(None, map(str, epic_resource_list)))))
    epics_df.loc[epics_df['E_key'] == epic_id, 'resource'] = epic_resource

# Step 7: Assign resource value of initiatives
unique_initiative_ids = initiatives_df['I_key'].unique()
for initiative_id in unique_initiative_ids:
    initiative_row = initiatives_df[initiatives_df['I_key'] == initiative_id].iloc[0]
    initiative_epics = epics_df[epics_df['E_ParentLink'] == initiative_id]
    initiative_resource_list = [str(initiative_row['I_assignee'])] + str(initiative_row['I_AdditionalAssignee']).split(' | ') + list(initiative_epics['resource'].astype(str))
    initiative_resource_list = [name for name in initiative_resource_list if str(name) != 'nan']  # Remove 'nan' values
    initiative_resource = ', '.join(list(set(filter(None, map(str, initiative_resource_list)))))
    initiatives_df.loc[initiatives_df['I_key'] == initiative_id, 'resource'] = initiative_resource

# Remove the intermediate resource_list column from stories_df
stories_df.drop('resource_list', axis=1, inplace=True)

# Step 8: Create the gantt_df dataframe
gantt_df = pd.DataFrame(columns=['issueType', 'issueKey', 'summary', 'TargetStart', 'TargetEnd', 'assignee', 'resources'])

for row in initiatives_df.itertuples():
    initiative = [row.I_key, row.I_summary, row.I_TargetStart, row.I_TargetEnd, row.I_assignee, row.resource]
    gantt_df.loc[len(gantt_df)] = ['Initiative', *initiative]

    initiative_epics = epics_df[epics_df['E_ParentLink'] == row.I_key]
    if len(initiative_epics) == 0:
        # If there are noepics associated with the initiative, add an empty row for the Epic section in the gantt_df.
```python
        gantt_df.loc[len(gantt_df)] = ['Epic', '', '', '', '', '', '']
    else:
        for epic_row in initiative_epics.itertuples():
            epic = [epic_row.E_key, epic_row.E_summary, epic_row.E_TargetStart, epic_row.E_TargetEnd, epic_row.E_assignee, epic_row.resource]
            gantt_df.loc[len(gantt_df)] = ['Epic', *epic]

            epic_stories = stories_df[stories_df['S_EpicLink'] == epic_row.E_key]
            if len(epic_stories) == 0:
                # If there are no stories associated with the epic, add an empty row for the Story section in the gantt_df.
                gantt_df.loc[len(gantt_df)] = ['Story', '', '', '', '', '', '']
            else:
                for story_row in epic_stories.itertuples():
                    story = [story_row.S_key, story_row.S_summary, story_row.S_TargetStart, story_row.S_TargetEnd, story_row.S_assignee, story_row.resource]
                    gantt_df.loc[len(gantt_df)] = ['Story', *story]

# Step 9: Save the gantt_df dataframe as a CSV file
gantt_df.to_csv('jira_gantt_all.csv', index=False)


import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Hyperlink

# Continue from the previous script
# ...

# Output the Gantt chart data as an Excel file
gantt_filename = 'gantt_chart.xlsx'
gantt_df.to_excel(gantt_filename, index=False)

# Modify the Gantt chart Excel file to create clickable entries
gantt_wb = pd.load_workbook(gantt_filename)
gantt_ws = gantt_wb.active

# Add hyperlink to each cell in the "issueKey" column
issueKey_col = gantt_df.columns.get_loc('issueKey') + 1
issueKey_col_letter = get_column_letter(issueKey_col)

for row_num, issueKey in enumerate(gantt_df['issueKey'], start=2):
    cell = gantt_ws[f'{issueKey_col_letter}{row_num}']
    cell.value = issueKey

    link = f'https://confluence.com/issue/{issueKey}'
    cell.hyperlink = Hyperlink(ref=f'{issueKey_col_letter}{row_num}', target=link)
    cell.font = Font(underline='single', color='0563C1')

# Save the modified Gantt chart Excel file
gantt_wb.save(gantt_filename)

# Additional steps with the modified Gantt chart Excel file
# ...
