import requests
import time
import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

SPLUNK_HOST = 'your_splunk_host'
SPLUNK_PORT = 8089
SPLUNK_USERNAME = 'your_username'
SPLUNK_PASSWORD = 'your_password'
SSL_CERT = "./HHHHTree.pem"
QUERY_DIR = './queries/'

create_job_url = f"https://{SPLUNK_HOST}:{SPLUNK_PORT}/services/search/jobs"


def run_query(query_file):
    with open(query_file, "r") as file:
        search_query = file.read()

    create_job_params = {
        "search": f"search {search_query}",
        "exec_mode": "normal",
        "earliest_time": "0",
        "latest_time": "now",
    }

    response = requests.post(
        create_job_url,
        auth=(SPLUNK_USERNAME, SPLUNK_PASSWORD),
        data=create_job_params,
        verify=SSL_CERT
    )

    sid_start = response.text.find("<sid>") + 5
    sid_end = response.text.find("</sid>")
    sid = response.text[sid_start:sid_end]

    job_status_url = f"{create_job_url}/{sid}"
    job_is_done = False

    while not job_is_done:
        response = requests.get(
            job_status_url,
            auth=(SPLUNK_USERNAME, SPLUNK_PASSWORD),
            verify=SSL_CERT
        )

        is_done_start = response.text.find("<isDone>") + 8
        is_done_end = response.text.find("</isDone>")
        is_done_value = response.text[is_done_start:is_done_end]
        job_is_done = is_done_value == "1"
        time.sleep(2)

    results_url = f"{job_status_url}/results"
    results_params = {
        "output_mode": "csv",
    }

    response = requests.get(
        results_url,
        auth=(SPLUNK_USERNAME, SPLUNK_PASSWORD),
        params=results_params,
        verify=SSL_CERT
    )

    output_file = f"{os.path.splitext(query_file)[0]}.csv"
    with open(output_file, 'wb') as f:
        f.write(response.content)

    print(f"CSV results saved to {output_file}")
    return output_file


def csv_to_xlsx(csv_file):
    df = pd.read_csv(csv_file)
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        length = min(length, 60)  # Maximum width = 60
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    xlsx_file = f"{os.path.splitext(csv_file)[0]}.xlsx"
    wb.save(xlsx_file)
    print(f"Excel file saved to {xlsx_file}")


query_files = glob.glob(f"{QUERY_DIR}/*.txt")
for query_file in query_files:
    csv_file = run_query(query_file)
    csv_to_xlsx(csv_file)
