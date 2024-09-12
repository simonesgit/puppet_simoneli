import os
import requests
from openpyxl import Workbook

SPLUNK_HOST = 'localhost'   
SPLUNK_PORT = 8089                 
SPLUNK_USERNAME = 'admin'
SPLUNK_PASSWORD = 'changeme'

QUERIES_FOLDER = './queries'

SSL_CERT = "./HHHHTree.pem"

def run_search_and_save(search_query, name):
    
    response = requests.post(f"https://{SPLUNK_HOST}:{SPLUNK_PORT}/services/search/jobs",
        auth=(SPLUNK_USERNAME, SPLUNK_PASSWORD),
        json={"search": search_query},
        verify=SSL_CERT
    )

    sid = response.json()["sid"]

    job_url = f"https://{SPLUNK_HOST}:{SPLUNK_PORT}/services/search/jobs/{sid}"
    while True:
        response = requests.get(job_url, auth=(SPLUNK_USERNAME, SPLUNK_PASSWORD), verify=SSL_CERT)
        status = response.json()["entry"][0]["content"]["state"]
        if status not in ["running", "pending"]:
            break
        
   results_url = f"{job_url}/results"     
   response = requests.get(results_url, auth=(SPLUNK_USERNAME, SPLUNK_PASSWORD), params={"output_mode": "csv"}, verify=SSL_CERT)
    
   filename = f"results/{name}.csv"
   with open(filename, "wb") as f:
       f.write(response.content)
       
   return sid, filename
   
def import_csv_to_sheet(filename, worksheet):
   with open(filename) as f:
       for r, row in enumerate(f, 1):
           for c, value in enumerate(row.split(", "), 1):
               worksheet.cell(row=r, column=c).value = value
               
def adjust_column_widths(workbook):
   for worksheet in workbook:
       for column in worksheet.columns:
           max_length = 0
           for cell in column:
               try:
                   if len(str(cell.value)) > max_length:
                       max_length = len(str(cell.value))
               except:
                   pass
           max_length = min(max_length, 60)  
           worksheet.column_dimensions[column[0].column_letter].width = max_length
           
# Main            
wb = Workbook()

for query_file in os.listdir(QUERIES_FOLDER):
    
    query_name = os.path.splitext(query_file)[0]
    
    with open(os.path.join(QUERIES_FOLDER, query_file)) as f:
        search_query = f.read()
        
    sid, results_file = run_search_and_save(search_query, query_name)
    
    ws = wb.create_sheet(title=query_name)  
    import_csv_to_sheet(results_file, ws)
    
adjust_column_widths(wb)

wb.save('output.xlsx')
